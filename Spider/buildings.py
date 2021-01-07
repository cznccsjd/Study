#coding:utf-8
"""
获取楼盘信息的公共方法
"""
import os, requests, sys, time
from bs4 import BeautifulSoup
from openpyxl import Workbook

class Building:
    #todo 新增一个excel
    def add_excel_for_house_detail(self):
        pass

    def file_is_exists(self, filename):
        return os.path.exists(filename)

    def get_sign_contract_infos(self, url):
        if not url:
            print("url不能为空")
            sys.exit()

        req = requests.get(url)
        res = req.text

        soup = BeautifulSoup(res, 'html.parser')
        divs = soup.find_all('div', {'id':'Div_presellItemMortagageInfo'})
        sign_infos = divs[2].text

        if "期房签约统计" not in sign_infos:
            print("目前没有期房签约统计数据了，排查下官网吧")
            return False

        if "住宅" not in sign_infos:
            print("目前没有住宅数据，后续再来查询吧")
            return False

        info_tmps = sign_infos.split('住宅')[1]
        info_tmp = info_tmps.split('本网站')[0]

        if "丙类储藏间" in info_tmp:
            info_houses = info_tmp.split('丙类储藏间')[0]
            info_stores = info_tmp.split('丙类储藏间')[1]

            houses = info_houses.split('\n')
            stores = info_stores.split('\n')

            # 住宅已签约套数
            sign_houses_num_total = houses[1]
            # 住宅已签约面积
            sign_houses_area_total = houses[2]
            # 住宅成交均价
            sign_houses_price_avg = houses[3]
            # 丙类储物间已签约套数
            sign_stores_num_total = stores[1]
            # 丙类储物间已签约面积
            sign_stores_area_total = stores[2]
            # 丙类储物间成交均价
            sign_stores_price_avg = stores[3]

        return [sign_houses_num_total, sign_houses_area_total, sign_houses_price_avg, sign_stores_num_total, sign_stores_area_total, sign_stores_price_avg]

    # 确认打开的页面是否正确
    def verify_page_infos(self, url, project_name):
        req = requests.get(url)
        res = req.text

        soup = BeautifulSoup(res, 'html.parser')
        # 先确认是项目名称是否正确
        project_infos = soup.find('td', {'id':'项目名称'})
        if project_name not in project_infos.contents[0]:
            print("获取的不是%s项目" % project_name)
            sys.exit()

        # 判断楼盘表格是否存在
        building_title = soup.find('td', {'class':'font_title2'})
        if "楼盘表" not in building_title.contents[2]:
            print("没有获取到楼盘表信息")
            sys.exit()


    def test(self):
        # excel相关操作
        wb = Workbook()
        ws = wb.active
        dir = os.getcwd()
        filetime = time.strftime('_%Y%m%d_%H%M%S', time.localtime())
        file = 'sign_contract_statistics.xlsx'
        # 绘制Excel表格
        ws.cell(row=1, column=1).value = '楼号'
        ws.cell(row=1, column=2).value = '门牌号'
        ws.cell(row=1, column=3).value = '户型'
        ws.cell(row=1, column=4).value = '建筑面积'
        ws.cell(row=1, column=5).value = '套内面积'
        ws.cell(row=1, column=6).value = '按建筑面积拟售单价'
        ws.cell(row=1, column=7).value = '按套内面积拟售单价'
        ws.cell(row=1, column=8).value = '分摊面积'
        ws.cell(row=1, column=9).value = '建筑面积总价'

        # {'2':{'1-501':['两室一厅','61.8500','50.3300','45094.93','55416.68']}}
        for key, value in datas.items():
            for i in value:
                ws_max_row = ws.max_row
                ws_max_col = ws.max_column
                ws.cell(row=ws_max_row+1, column=1).value = key
                ws.cell(row=ws_max_row+1, column=2).value = i
                ws.cell(row=ws_max_row+1, column=3).value = value[i][0]
                ws.cell(row=ws_max_row+1, column=4).value = value[i][1]
                ws.cell(row=ws_max_row+1, column=5).value = value[i][2]
                ws.cell(row=ws_max_row+1, column=6).value = value[i][3]
                ws.cell(row=ws_max_row+1, column=7).value = value[i][4]
                linenums = ws_max_row

        # 保存Excel（可以覆盖保存）
        wb.save(file)
        print('excel数据写入成功：',time.strftime('%H:%M:%S',time.localtime()))
        print('文件名：',file)
        print('写入%d条数据' % linenums)
        print('脚本结束时间:', time.strftime('%H:%M:%S',time.localtime()))

if __name__ == '__main__':
    b = Building()
    aa = b.file_is_exists('jingxiyinyue_20201231_021842.xlsx')
    print(aa)
    print("he")