#coding:utf-8
#房屋价格信息

import os, random, re, requests, sys, time
from bs4 import BeautifulSoup
from openpyxl import Workbook
from Spider.buildings import Building

# 楼盘的正则条件
#楼号
build_num = '\d{2}-\d+'
#本楼可销售数量
build_sale_avail_num = '\n\d{2}\n'
# 本楼均价
build_sale_ave_price = '\d{5}.\d{2}'

def get_building_detail(datas=None):
    detail_infos = {}
    header = {'Connection': 'close'}

    for key, value in datas.items():
        louhao = key
        url = value[2]
        dict_tmp = {}

        time.sleep(1)
        reponse = requests.get(url, headers=header)
        res = reponse.text

        soup = BeautifulSoup(res, 'html.parser')
        divs = soup.find_all('tbody')
        house_infos_tmps = divs[10]
        house_lists_tmps = house_infos_tmps.find_all('div')
        for house_lists_tmp in house_lists_tmps:
            # 具体房间的url
            house_url = house_lists_tmp.contents[1].attrs['href']
            # 具体的楼号
            name_tmp = house_lists_tmp.text
            danyuan_haoma = re.sub('\n', '', name_tmp)

            t = random.randint(2,5)
            time.sleep(t)
            res_house_r = requests.get('http://bjjs.zjw.beijing.gov.cn/' + house_url, headers=header)
            res_house = res_house_r.text
            soup_house = BeautifulSoup(res_house, 'html.parser')
            list_tmps = soup_house.find('div',{'id':'showDiv'})
            h_infos = list_tmps.find_all('tr')

            # 户型
            huxing_tmp = h_infos[3].text
            m = re.search('[\u4e00-\u9fa5]{4}', huxing_tmp)
            if m is not None:
                huxing = m.group()

            # 建筑面积
            jianzhu_mianji_tmp = h_infos[4].text
            m = re.search('\d+.\d+', jianzhu_mianji_tmp)
            if m is not None:
                jianzhu_mianji = m.group()

            # 套内面积
            taonei_mianji_tmp = h_infos[5].text
            m = re.search('\d+.\d+', taonei_mianji_tmp)
            if m is not None:
                taonei_mianji = m.group()

            # 按建筑面积拟售单价
            jianzhu_danjia_tmp = h_infos[6].text
            m = re.search('\d+.\d+', jianzhu_danjia_tmp)
            if m is not None:
                jianzhu_danjia = m.group()

            # 按套内面积拟售单价
            taonei_danjia_tmp = h_infos[7].text
            m = re.search('\d+.\d+', taonei_danjia_tmp)
            if m is not None:
                taonei_danjia = m.group()

            #存储的格式 {'2':{'1-501':['两室一厅','61.8500','50.3300','45094.93','55416.68']}}
            dict_tmp[danyuan_haoma] = [huxing, jianzhu_mianji, taonei_mianji, jianzhu_danjia, taonei_danjia]
        detail_infos[louhao] = dict_tmp

    print(detail_infos)
    set_excel(detail_infos)

# 获取楼盘概要
def get_building_summary():
    '''
    获取楼盘概要
    :return: {'销售楼名':['批准销售套数','预售住宅拟售均价','该楼栋详细信息链接']}
    '''
    print('脚本开始运行:', time.strftime('%H:%M:%S',time.localtime()))
    build_summary_infos = {}
    # 京西印玥
    url = 'http://bjjs.zjw.beijing.gov.cn/eportal/ui?pageId=411617&systemId=2&srcId=1&id=6849021&rowcount=21'

    req = requests.get(url)
    res =req.text

    soup = BeautifulSoup(res, 'html.parser')
    # 先确认是兴辰佳苑
    project_infos = soup.find('td', {'id':'项目名称'})
    if "兴辰佳苑" not in project_infos.contents[0]:
        print("获取的不是兴辰佳苑[中建.京西印玥]项目")
        sys.exit()

    # 判断楼盘表格是否存在
    building_title = soup.find('td', {'class':'font_title2'})
    if "楼盘表" not in building_title.contents[2]:
        print("没有获取到楼盘表信息")
        sys.exit()

    building_divs = soup.find('form', {'id':'FDCJYFORM'})
    building_lists = building_divs.find_all('tr',{'align':'center'})
    for building_list in building_lists:
        building_infos = building_list.text
        if "住宅楼" not in building_infos:
            continue

        m = re.search(build_num, building_infos)
        if (m is not None):
            b_nums = m.group()
            b_num = b_nums.split('-')[1]

        m = re.search(build_sale_avail_num, building_infos)
        if(m is not None):
            b_sale_avail_nums = m.group()
            b_sale_avail_num = re.sub('\n','',b_sale_avail_nums)

        m = re.search(build_sale_ave_price, building_infos)
        if(m is not None):
            b_sale_ave_price = m.group()

        # 获取每栋楼的url
        urls = building_list.contents[1].contents[1].attrs['href']
        url = re.sub('\r|\n|\t','',urls)

        build_summary_infos[b_num] = [b_sale_avail_num, b_sale_ave_price, 'http://bjjs.zjw.beijing.gov.cn/'+url]
    print(build_summary_infos)
    return build_summary_infos

# 获取签约信息
def get_sign_contract():
    url = 'http://bjjs.zjw.beijing.gov.cn/eportal/ui?pageId=320801&projectID=6849021&systemID=2&srcId=1'
    print("京西印玥截止到今天的签约数据如下：\n")
    ret = Building().get_sign_contract_infos(url)
    if ret:
        print(ret)

def set_excel(datas):
    # excel相关操作
    wb = Workbook()
    ws = wb.active
    dir = os.getcwd()
    filetime = time.strftime('_%Y%m%d_%H%M%S', time.localtime())
    file = 'jingxiyinyue' + filetime + '.xlsx'
    # 绘制Excel表格
    ws.cell(row=1, column=1).value = '楼号'
    ws.cell(row=1, column=2).value = '门牌号'
    ws.cell(row=1, column=3).value = '户型'
    ws.cell(row=1, column=4).value = '建筑面积'
    ws.cell(row=1, column=5).value = '套内面积'
    ws.cell(row=1, column=6).value = '按建筑面积拟售单价'
    ws.cell(row=1, column=7).value = '按套内面积拟售单价'
    ws.cell(row=1, column=8).value = '分摊面积'
    ws.cell(row=1, column=9).value = '建筑面积总价'

    # {'2':{'1-501':['两室一厅','61.8500','50.3300','45094.93','55416.68']}}
    for key, value in datas.items():
        for i in value:
            ws_max_row = ws.max_row
            ws_max_col = ws.max_column
            ws.cell(row=ws_max_row+1, column=1).value = key
            ws.cell(row=ws_max_row+1, column=2).value = i
            ws.cell(row=ws_max_row+1, column=3).value = value[i][0]
            ws.cell(row=ws_max_row+1, column=4).value = value[i][1]
            ws.cell(row=ws_max_row+1, column=5).value = value[i][2]
            ws.cell(row=ws_max_row+1, column=6).value = value[i][3]
            ws.cell(row=ws_max_row+1, column=7).value = value[i][4]
            linenums = ws_max_row

    # 保存Excel（可以覆盖保存）
    wb.save(file)
    print('excel数据写入成功：',time.strftime('%H:%M:%S',time.localtime()))
    print('文件名：',file)
    print('写入%d条数据' % linenums)
    print('脚本结束时间:', time.strftime('%H:%M:%S',time.localtime()))


# infos = get_building_summary()
# get_building_detail(infos)
get_sign_contract()