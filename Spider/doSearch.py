#coding:utf-8
'''
执行豆瓣爬虫文件
'''
import time, datetime, requests, os,re
from Spider import douban
from bs4 import BeautifulSoup
from openpyxl import Workbook

class DoSearch():
    def __init__(self):
        #本脚本开始执行时间，方便需要时查看执行效率
        print('脚本开始运行:', time.strftime('%H:%M:%S',time.localtime()))
        self.engine = douban.Douban()
        self.douB = self.engine.search()
        self.configDouban = self.engine.config

    def doDouBan(self):
        '''
        从douban.py获取url，执行爬虫操作
        后续需要拆分出多个方法
        :return:
        '''
        sleeptime = self.configDouban['sleep']
        wantedAreas = self.configDouban['Area']
        otherfilter = self.configDouban['titlelimit']
        wantedAreasStr = self.engine.restr(wantedAreas)
        otherfilter = self.engine.restr(otherfilter)
        errorNum = self.configDouban['errorNum']
        preTime = self.ftime
        linenums = 0

        # excel相关操作
        wb = Workbook()
        ws = wb.active
        dir = os.getcwd()
        filetime = time.strftime('_%Y%m%d_%H%M%S', time.localtime())
        file = 'douban' + filetime + '.xlsx'
        # 绘制Excel表格
        ws.cell(row=1, column=1).value = '标题'
        ws.cell(row=1, column=2).value = 'URL'
        ws.cell(row=1, column=3).value = '价格'
        ws.cell(row=1, column=4).value = '地区'
        ws.cell(row=1, column=5).value = '最后更新时间'
        ws.cell(row=1, column=6).value = '来源'
        ws.cell(row=1, column=7).value = '微信'
        ws.cell(row=1, column=8).value = 'TEL'
        ws.cell(row=1, column=9).value = '户型'

        #打开单个组，获取该组下所有有价值的单条url
        for url in self.douB.keys():
            urlnew = url + 'discussion'
            pagetitle = self.douB[url]
            maxtitles = self.configDouban['titlemaxmun']
            maxtitle = int(maxtitles)
            if maxtitle <= 2000:
                pass
            else:
                maxtitle = 2000     #人为设置2000，最多获取1975页

            for num in range (0,maxtitle,25):
                data = {'start': num}
                header = self.configDouban['headers']
                cookiea = self.configDouban['cookie']
                # 关闭多余的连接
                s = requests.sessions.session()
                s.keep_alive = False

                try:
                    res = requests.get(urlnew,data, headers = header, cookies = cookiea,timeout=10)
                    # 增加request间隔时间，防止被封IP
                    time.sleep(sleeptime)
                except Exception as e:
                    errorNum += 1
                    self.configDouban['errorNum'] = errorNum
                    print("%s,第%d次报错，url：%s,原因：%s" %(time.strftime('%H:%M:%S',time.localtime()),errorNum,urlnew,e))
                    continue

                soup = BeautifulSoup(res.text,"lxml")
                titles = soup.select('td.title > a')
                links = soup.select('td.title > a')
                times = soup.select('td.time')

                for title, link, atime in zip(titles, links, times):
                    title = title.text
                    link = link.get('href')
                    atime = atime.text
                    # 当前这条数据的时间在5天前，就不要了，并且终止后面的循环(for else)
                    if atime < preTime:
                        break
                    # 通过标题查找符合要求的地区
                    filterPagetitle = re.search(wantedAreasStr, title)
                    if filterPagetitle is None:
                        continue
                    # 标题含有已租、限女的，也要去除
                    filterPagetitlesex = re.search(otherfilter, title)
                    if filterPagetitlesex is not None:
                        continue

                    # 打开具体某一条招租信息
                    try:
                        s = requests.sessions.session()
                        s.keep_alive = False
                        rs = requests.get(link, headers=header, cookies=cookiea, timeout=10)
                        # 增加request间隔时间，防止被封IP
                        time.sleep(sleeptime)
                    except Exception as e:
                        errorNum += 1
                        self.configDouban['errorNum'] = errorNum
                        print("%s,第%d次报错，url：%s,原因：%s" % (
                            time.strftime('%H:%M:%S', time.localtime()), errorNum, link, e))
                        continue
                    soup = BeautifulSoup(rs.text, "lxml")

                    if '...' in title:
                        stitle1 = soup.select('td.tablecc')
                        stitle2 = str(stitle1)
                        stitle = stitle2[41:-6]
                    else:
                        stitle = title

                    filtertitle = re.search(wantedAreasStr, stitle)
                    if filtertitle is None:
                        continue

                    filtertitle1 = re.search(otherfilter, stitle)
                    if filtertitle1 is not None:
                        continue

                    contexts = soup.select('div.topic-richtext')
                    context = str(contexts)
                    conIndexsBegin = context.find('<p>')
                    conIndexsEnd = context.find('</p><div class="image-container image-float-center">')
                    context = context[conIndexsBegin:conIndexsEnd + 1]

                    if context == '' or context == '<p><':
                        continue
                    elif stitle == '':
                        continue
                    # 正则表达式，筛选出金额
                    patMoney = '[\D][\d]{4}[\D]'
                    # patTel = '[\D]?[\d]{11}[\D]?'
                    # patWe = '[微信]|[wechat]'
                    moneyTitle = re.findall(patMoney, stitle)
                    moneyContext = re.findall(patMoney, context)
                    smoney1 = moneyTitle + moneyContext
                    smoney2 = str(smoney1)        #openpyxl将str写入一个单元格
                    smoney = re.sub('\[|\]|\'','',smoney2)
                    # stel = re.search(patTel,context)
                    # swechat = re.search(patWe,context)

                    ws_max_row = ws.max_row
                    ws_max_col = ws.max_column
                    ws.cell(row=ws_max_row+1, column=1).value = stitle.strip()
                    ws.cell(row=ws_max_row+1, column=5).value = atime
                    ws.cell(row=ws_max_row+1, column=2).value = link
                    ws.cell(row=ws_max_row+1, column=6).value = pagetitle
                    ws.cell(row=ws_max_row+1, column=3).value = smoney
                    linenums = ws_max_row
                    continue
                else:
                    continue
                break

        # 保存Excel（可以覆盖保存）
        wb.save(file)
        print('excel数据写入成功：',time.strftime('%H:%M:%S',time.localtime()))
        print('文件名：',file)
        print('写入%d条数据' % linenums)

    @property
    def ftime(self):
        # 获取指定x天前时间
        day = self.configDouban['date']
        day = int(day)
        preday = (datetime.datetime.now() - datetime.timedelta(days=day)).strftime("%m-%d %H:%M")
        return preday

if __name__ == '__main__':
    DoSearch().doDouBan()
    # DoSearch().ftime()
    # DoSearch().config()
    # DoSearch().excel()