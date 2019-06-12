#coding:utf-8
'''
简单学习Openpyxl文件
'''

import openpyxl

file = r'D:\Documents\work\工作记录\180523 缺席代课报表\缺席代课报表数据比对.xlsx'

######################################
#   workbook对象      #
######################################
# 新建workbook对象
# wb = openpyxl.Workbook()
# print('新建的Workbook对象:', wb)

# 打开已存在的Excel
wb = openpyxl.load_workbook(file)

# 已列表的形式返回所有的Worksheet（表格）
wss = wb.worksheets

###### 获取表格信息 ######
# 获取当前活跃的WorkSheet
ws = wb.active

# 判断是否以read_only模式打开Excel文档
ws_read = wb.read_only

# 获取文档的字符集编码
ws_encode = wb.encoding

# 获取文档的元数据，如标题、创建者、创建日期等
ws_pro = wb.properties

# 获取工作簿中的表
ws_sheet = wb.sheetnames

###### 新增、删除操作 ######
# 创建一个空的表格
wb.create_sheet

# 在Workbook内拷贝表格
'''
copy_worksheet() 需要的参数是某个活跃的worksheet       
后续补充打开查看excel文件内容，验证只是拷贝了sheet名称，还是也拷贝了sheet内容
'''
wb.copy_worksheet(ws)

# 删除一个表格 remove_sheet()
wss = wb.worksheets
wb.remove_sheet(wss[4])




######################################
#   worksheet对象      #
######################################
# 表格的标题
ws_title = ws.title

# 表格的大小，这里的大小是指含有数据的表格的大小，即：左上角的坐标，右下角的坐标
ws_dimensions = ws.dimensions

# 表格的最大行
ws_max_row = ws.max_row

# 表格的最小行
ws_min_row = ws.min_row

# 表格的最大列
ws_max_column = ws.max_column

# 表格最小列
ws_min_column = ws.min_column

# 按行获取单元格（Cell对象）-生成器
ws_rows = ws.rows

# 按列获取单元格（Cell对象）-生成器
ws_columns = ws.columns

# 冻结窗格
"""
freeze_panes，参数比较特别，主要用于在表格较大时冻结顶部的行或左边的行。对于冻结的行，在用户滚动时，是始终可见的，
可以设置为一个Cell对象或一个端元个坐标的字符串，单元格上面的行和左边的列将会冻结(单元格所在的行和列不会被冻结)。
例如我们要冻结第一行那么设置A2为freeze_panes，如果要冻结第一列，freeze_panes取值为B1，如果要同时冻结第一行和第一列，
那么需要设置B2为freeze_panes，freeze_panes值为none时 表示 不冻结任何列。
"""
ws_freeze_panes = ws.freeze_panes

# 按行获取表格的内容（数据）-生成器
ws_values = ws.values

######### 常用的worksheet方法如下
# 按行获取所有单元格，内置属性有（min_row,max_row,min_col,max_col）
ws_iter_rows = ws.iter_rows

# 按列获取所有的单元格    提示：AttributeError: 'Worksheet' object has no attribute 'iter_columns'
ws_iter_columns = ws.iter_columns

# 在表格末尾添加数据
ws_append = ws.append

# 合并多个单元格
ws_merged_cells = ws.merged_cells

# 移除合并的单元格  运行报错“AttributeError: 'Worksheet' object has no attribute 'unmerged_cells'”
ws_unmerged_cells = ws.unmerged_cells




######################################
#   Cell对象      #常用的属性如下
######################################
'''
单元格所在的行   row
单元格所在的列   column
单元格的值     value
单元格的坐标        coordinate
'''
# 单元格的坐标
ws_coordinate = ws.cell(row=1, column=2).coordinate

# 单元格的值
ws_value = ws.cell(row=1, column=2).value

# 单元格所在的行
ws_row = ws.cell(row=1, column=2).row

# 单元格所在的列
ws_column = ws.cell(row=1, column=2).column

print(ws_rows)